# app/routes/api.py
import csv
import io
import os
import requests
import uuid
import decimal
from datetime import datetime, date, timedelta
from flask import Blueprint, jsonify, request, session, send_file
from functools import wraps
from app import mysql
from app.utils import calculate_ean13_checksum
from .auth import login_required
import openpyxl
import json
import pyodbc

api_bp = Blueprint('api', __name__)
def get_sql_server_connection():
    """Membuat dan mengembalikan koneksi ke database SQL Server."""
    try:
        conn = pyodbc.connect(
            'DRIVER={' + os.getenv('SQL_SERVER_DRIVER') + '};'
            'SERVER=' + os.getenv('SQL_SERVER_HOST') + ';'
            'DATABASE=' + os.getenv('SQL_SERVER_DB') + ';'
            'UID=' + os.getenv('SQL_SERVER_USER') + ';'
            'PWD=' + os.getenv('SQL_SERVER_PASSWORD') + ';'
        )
        return conn
    except pyodbc.Error as ex:
        # Log error atau tangani sesuai kebutuhan
        print(f"Gagal terhubung ke SQL Server: {ex}")
        return None

# --- Decorator untuk Izin Akses ---
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            permissions = session.get('permissions', [])
            if permission not in permissions:
                return jsonify({"status": "error", "message": "Akses ditolak. Anda tidak memiliki izin."}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

permission_map = {
    'items': {'R': 'RI', 'W': 'WI', 'D': 'DI'}, 
    'penjualan': {'R': 'RTk'},
    'materials': {'R': 'RM', 'W': 'WM', 'D': 'DM'}, 
    'sizes': {'R': 'RSz', 'W': 'WSz', 'D': 'DSz'},
    'brands': {'R': 'RB', 'W': 'WB', 'D': 'DB'}, 
    'models': {'R': 'RM', 'W': 'WM', 'D': 'DM'},
    # --- PERUBAHAN DI SINI ---
    'penerimaan': {'R': 'RTk'},
    'piutang': {'R': 'RTk'},
    'analytics': {'R': 'RA'},
    'pajak-invoice': {'R': 'RPjk'},
    'factory-analytics': {'R': 'RA'},
    'multipayroll': {'R': 'RA'},
    'payroll-checksum': {'R': 'RA'},
    'packinglist-barcode': {'R': 'RPL', 'W': 'WPL'},
    'packinglist-riwayat': {'R': 'RPL'},
    'gudang': {'R': 'RGudang'},
    'stock-in': {'R': 'RGudang', 'W': 'WStok'},
    'stock-out': {'R': 'RGudang', 'W': 'WStok'}
}

# --- Rute-rute API ---

@api_bp.route('/user/permissions')
@login_required
def get_user_permissions():
    return jsonify({"status": "success", "permissions": session.get('permissions', [])})

@api_bp.route('/data/<data_type>', methods=['GET'])
@login_required
def get_data(data_type):
    perm_code = permission_map.get(data_type, {}).get('R')
    if not perm_code or perm_code not in session.get('permissions', []):
        return jsonify({"status": "error", "message": "Akses ditolak."}), 403
    
    search_query = request.args.get('search', '')
    sort_by = request.args.get('sort_by', 'CreatedDate')
    sort_order = request.args.get('sort_order', 'DESC')
    is_super_admin = 'SA' in session.get('permissions', [])

    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    offset = (page - 1) * limit

    if data_type in ['packinglist-barcode', 'packinglist-riwayat']:
        conn_sql = get_sql_server_connection()
        if not conn_sql:
            return jsonify({"status": "error", "message": "Gagal terhubung ke database Packing List."}), 503
        
        try:
            cursor = conn_sql.cursor()
            params = []
            where_clause = ""
            
            if data_type == 'packinglist-barcode':
                query_base = "FROM dbo.ViewBarcodeSystemCMT"
                query_fields = "Barcode, Item, Color, Size"
                order_by_clause = "ORDER BY Barcode"
                count_field = "Barcode"
                if search_query:
                    where_clause = " WHERE Barcode LIKE ? OR Item LIKE ? OR Size LIKE ?"
                    params.extend([f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'])
            
            elif data_type == 'packinglist-riwayat':
                query_base = """
                    FROM dbo.PackListCmt p
                    LEFT JOIN dbo.ViewProdNoCustomerExt c ON p.CustomerID = c.CustomerID
                """
                query_fields = "p.IDNo, c.Customer AS CustomerName, p.DONo, p.WONo"
                # --- PERBAIKAN DI SINI: Hapus alias 'p.' dari ORDER BY ---
                order_by_clause = "ORDER BY IDNo DESC"
                # --- AKHIR PERBAIKAN ---
                count_field = "DISTINCT p.IDNo"
                if search_query:
                    where_clause = " WHERE p.WONo LIKE ? OR p.DONo LIKE ? OR c.Customer LIKE ?"
                    params.extend([f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'])

            # 1. Query hitung total baris (tidak berubah)
            count_query = f"SELECT COUNT({count_field}) {query_base} {where_clause}"
            cursor.execute(count_query, params if search_query else [])
            total_records = cursor.fetchone()[0]

            # 2. Query untuk mendapatkan data menggunakan ROW_NUMBER()
            start_row = offset + 1
            end_row = offset + limit
            
            if data_type == 'packinglist-riwayat':
                 final_fields = "IDNo, CustomerName, DONo, WONo" # Nama kolom setelah alias
                 query = f"""
                    WITH DistinctResults AS (
                        SELECT DISTINCT {query_fields}
                        {query_base} {where_clause}
                    )
                    , NumberedResults AS (
                        SELECT *, ROW_NUMBER() OVER ({order_by_clause}) as row_num
                        FROM DistinctResults
                    )
                    SELECT {final_fields}
                    FROM NumberedResults
                    WHERE row_num BETWEEN ? AND ?;
                """
            else: # Untuk barcode, query lebih sederhana
                final_fields = query_fields
                query = f"""
                    WITH NumberedResults AS (
                        SELECT {query_fields}, ROW_NUMBER() OVER ({order_by_clause}) as row_num
                        {query_base} {where_clause}
                    )
                    SELECT {final_fields}
                    FROM NumberedResults
                    WHERE row_num BETWEEN ? AND ?;
                """

            params.extend([start_row, end_row])
            cursor.execute(query, params)
            
            columns = [column[0] for column in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            return jsonify({"status": "success", "data": results, "total_records": total_records})

        except pyodbc.Error as ex:
            return jsonify({"status": "error", "message": f"Database error: {str(ex)}"}), 500
        finally:
            if conn_sql:
                conn_sql.close()

    elif data_type == 'pajak-invoice':
        conn_sql = get_jbbdata_sql_server_connection()
        if not conn_sql:
            return jsonify({"status": "error", "message": "Gagal terhubung ke database JBBData."}), 503
        
        try:
            cursor = conn_sql.cursor()
            params = []
            
            query_base = "FROM dbo.ViewSalInvDetPajakExt"
            query_fields = "IDNo, InvNo, CONVERT(VARCHAR, InvDate, 105) as InvDate, Toko, Customer, Qty, Up, Amount, Disc, TotalAmount, DppUp, DppDisc, DppAmt, DppNilaiLain, VatAmount, Remark"
            count_field = "IDNo"
            
            valid_sort_columns = ['IDNo', 'InvNo', 'InvDate', 'Toko', 'Customer', 'Qty', 'Up', 'Amount', 'Disc', 'TotalAmount', 'DppUp', 'DppDisc', 'DppAmt', 'DppNilaiLain', 'VatAmount', 'Remark']
            if sort_by not in valid_sort_columns:
                sort_by = 'InvDate'
            if sort_order.upper() not in ['ASC', 'DESC']:
                sort_order = 'DESC'
            
            order_by_clause = f"ORDER BY {sort_by} {sort_order}"

            where_clauses = []
            
            if search_query:
                where_clauses.append("(InvNo LIKE ? OR Toko LIKE ? OR Customer LIKE ?)")
                search_term = f'%{search_query}%'
                params.extend([search_term, search_term, search_term])
            
            toko = request.args.get('toko')
            if toko:
                where_clauses.append("Toko = ?")
                params.append(toko)

            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if start_date and end_date:
                where_clauses.append("CAST(InvDate AS DATE) BETWEEN ? AND ?")
                params.extend([start_date, end_date])

            where_clause_str = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
            
            count_query = f"SELECT COUNT({count_field}) {query_base} {where_clause_str}"
            cursor.execute(count_query, params)
            total_records = cursor.fetchone()[0]

            start_row = offset + 1
            end_row = offset + limit
            
            query = f"""
                WITH NumberedResults AS (
                    SELECT {query_fields}, ROW_NUMBER() OVER ({order_by_clause}) as row_num
                    {query_base} {where_clause_str}
                )
                SELECT * FROM NumberedResults WHERE row_num BETWEEN ? AND ?;
            """
            
            final_params = params + [start_row, end_row]
            cursor.execute(query, final_params)
            
            columns = [column[0] for column in cursor.description if column[0] != 'row_num']
            results = [dict(zip(columns, [str(v) if isinstance(v, decimal.Decimal) else v for v in row])) for row in cursor.fetchall()]
            
            return jsonify({"status": "success", "data": results, "total_records": total_records})

        except pyodbc.Error as ex:
            return jsonify({"status": "error", "message": f"Database error: {str(ex)}"}), 500
        finally:
            if conn_sql: conn_sql.close()

    if sort_order.upper() not in ['ASC', 'DESC']:
        sort_order = 'DESC'
    try:
        cur = mysql.connection.cursor()
        params = []
        
        base_query = ""
        count_query = ""
        
        if data_type == 'penjualan':
            valid_sort_columns = ['Nomor', 'Cabang', 'Tanggal', 'Salesman', 'Nama', 'Daerah', 'Telepon', 'TotalPenjualanLusin', 'TotalHargaPenjualan', 'Pembayaran', 'TanggalBayar', 'JatuhTempo']
            if sort_by not in valid_sort_columns:
                sort_by = 'a.CreatedDate'
            
            base_query = """
                FROM activities a
                LEFT JOIN branches b ON a.BranchId = b.Id
                LEFT JOIN salesmans s ON a.SalesmanId = s.Id
            """
            select_fields = """
                SELECT
                    a.Id,
                    DATE_FORMAT(a.Date, '%%y%%m%%d%%H%%i%%s') as Nomor, b.Name as Cabang, DATE_FORMAT(a.CreatedDate, '%%d-%%m-%%Y') as Tanggal,
                    s.Name as Salesman, a.Name as Nama, a.Address as Daerah, a.Phone as Telepon, a.Total as TotalPenjualanLusin,
                    a.TotalPrice as TotalHargaPenjualan, a.PaymentType as Pembayaran, DATE_FORMAT(a.PaymentDate, '%%d-%%m-%%Y') as TanggalBayar,
                    DATE_FORMAT(a.DueDate, '%%d-%%m-%%Y') as JatuhTempo, a.PaymentNote as KeteranganPembayaran,
                    CASE WHEN a.IsBypass = 1 THEN 'Ya' ELSE 'Tidak' END as Bypass,
                    CASE WHEN a.IsBypassVerified = 1 THEN 'Ya' ELSE 'Tidak' END as BypassDiterima
            """
            
            where_clauses = ["a.Type = 'SalesOrder'"]
            
            if search_query:
                where_clauses.append("(s.Name LIKE %s OR a.Name LIKE %s OR a.Phone LIKE %s OR b.Name LIKE %s)")
                search_term = f"%{search_query}%"
                params.extend([search_term, search_term, search_term, search_term])
            
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            if start_date and end_date:
                where_clauses.append("a.CreatedDate BETWEEN %s AND %s")
                params.extend([start_date, end_date + ' 23:59:59'])
            
            if where_clauses:
                where_clause_str = " WHERE " + " AND ".join(where_clauses)
                base_query += where_clause_str
            
            count_query = "SELECT COUNT(*) as total " + base_query
            query = select_fields + base_query + f" ORDER BY {sort_by} {sort_order} LIMIT %s OFFSET %s"

        elif data_type == 'penerimaan':
            valid_sort_columns = ['Nomor', 'Cabang', 'Dari', 'Tanggal', 'Scan', 'Keterangan', 'TotalPenjualanLusin', 'TotalHargaPenjualan']
            sort_by = request.args.get('sort_by', 'a.CreatedDate')
            if sort_by not in valid_sort_columns:
                sort_by = 'a.CreatedDate'

            base_query = """
                FROM activities a
                LEFT JOIN branches b ON a.BranchId = b.Id
                LEFT JOIN branches ref_b ON a.ReferenceBranchId = ref_b.Id
            """
            select_fields = """
                SELECT
                    DATE_FORMAT(a.Date, '%%y%%m%%d%%H%%i%%s') as Nomor,
                    b.Name as Cabang,
                    ref_b.Name as Dari,
                    DATE_FORMAT(a.CreatedDate, '%%d-%%m-%%Y') as Tanggal,
                    DATE_FORMAT(a.ScanDate, '%%d-%%m-%%Y %%H:%%i:%%s') as Scan,
                    a.Note as Keterangan,
                    a.Total as TotalPenjualanLusin,
                    a.TotalPrice as TotalHargaPenjualan
            """
            
            where_clauses = ["a.Type = 'Delivery'"]
            
            if search_query:
                where_clauses.append("(b.Name LIKE %s OR ref_b.Name LIKE %s OR a.Note LIKE %s)")
                search_term = f"%{search_query}%"
                params.extend([search_term, search_term, search_term])
            
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            if start_date and end_date:
                where_clauses.append("a.CreatedDate BETWEEN %s AND %s")
                params.extend([start_date, end_date + ' 23:59:59'])

            if where_clauses:
                base_query += " WHERE " + " AND ".join(where_clauses)
            
            count_query = "SELECT COUNT(*) as total " + base_query
            query = select_fields + base_query + f" ORDER BY {sort_by} {sort_order} LIMIT %s OFFSET %s"
        elif data_type == 'items':
            valid_sort_columns = ['Barcode', 'Name', 'BahanName', 'SizeName', 'BrandName', 'ModelName', 'Price', 'Serial', 'CreatedDate']
            if sort_by not in valid_sort_columns:
                sort_by = 'CreatedDate'
            
            column_mapping = {
                'BahanName': 'm.Name', 'SizeName': 's.Name', 'BrandName': 'b.Name', 'ModelName': 'mo.Name', 'Barcode': 'i.Barcode',
                'Name': 'i.Name', 'Price': 'i.Price', 'Serial': 'i.Serial', 'CreatedDate': 'i.CreatedDate'
            }
            db_sort_column = column_mapping.get(sort_by, 'i.CreatedDate')

            base_query = "FROM items i LEFT JOIN materials m ON i.MaterialId = m.Id LEFT JOIN sizes s ON i.SizeId = s.Id LEFT JOIN brands b ON i.BrandId = b.Id LEFT JOIN models mo ON i.ModelId = mo.Id"
            select_fields = "SELECT i.Id, i.Barcode, i.Name, i.Serial, m.Name as BahanName, s.Name as SizeName, b.Name as BrandName, mo.Name as ModelName, i.Price, i.IsActive, i.CreatedDate "
            
            where_clauses = []
            if search_query:
                where_clauses.append("(i.Name LIKE %s OR i.Barcode LIKE %s)")
                params.extend([f"%{search_query}%", f"%{search_query}%"])
            if not is_super_admin:
                where_clauses.append("i.IsActive = 1")
            
            if where_clauses:
                where_clause_str = " WHERE " + " AND ".join(where_clauses)
                base_query += where_clause_str
            
            count_query = "SELECT COUNT(*) as total " + base_query
            query = select_fields + base_query + f" ORDER BY {db_sort_column} {sort_order} LIMIT %s OFFSET %s"
        else:
            valid_sort_columns = ['Name', 'IsActive']
            if sort_by not in valid_sort_columns:
                sort_by = 'Name'
            base_query = f"FROM {data_type}"
            select_fields = f"SELECT Id, Name, IsActive "
            
            where_clauses = []
            if search_query:
                where_clauses.append("Name LIKE %s")
                params.append(f"%{search_query}%")
            if not is_super_admin:
                where_clauses.append("IsActive = 1")

            if where_clauses:
                where_clause_str = " WHERE " + " AND ".join(where_clauses)
                base_query += where_clause_str
                
            count_query = "SELECT COUNT(*) as total " + base_query
            query = select_fields + base_query + f" ORDER BY {sort_by} {sort_order} LIMIT %s OFFSET %s"

        cur.execute(count_query, tuple(params))
        total_records = cur.fetchone()['total']
        
        params.extend([limit, offset])
        cur.execute(query, tuple(params))
        results = cur.fetchall()
        cur.close()
        
        for row in results:
            for key, value in row.items():
                if isinstance(value, bytes):
                    row[key] = 1 if value == b'\x01' else 0
                elif isinstance(value, decimal.Decimal):
                    row[key] = str(value)
                elif isinstance(value, datetime):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({"status": "success", "data": results, "total_records": total_records})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/penjualan/header/<activity_id>')
@login_required
@permission_required('RP')
def get_penjualan_header(activity_id):
    try:
        cur = mysql.connection.cursor()
        query = """
            SELECT
                DATE_FORMAT(a.Date, '%%y%%m%%d%%H%%i%%s') as Nomor,
                a.Name,
                b.Name as Cabang,
                a.PaymentType as Pembayaran
            FROM activities a
            LEFT JOIN branches b ON a.BranchId = b.Id
            WHERE a.Id = %s
        """
        cur.execute(query, [activity_id])
        header_data = cur.fetchone()
        cur.close()
        return jsonify({"status": "success", "data": header_data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/penjualan/detail/<activity_id>')
@login_required
@permission_required('RP')
def get_penjualan_detail(activity_id):
    try:
        cur = mysql.connection.cursor()
        query = """
            SELECT
                i.Name AS 'NamaBarang',
                ad.Qty,
                ad.Price,
                ad.Discount,
                ad.Subtotal,
                ad.Note,
                ad.Serial,
                u.UserName as CreatedBy
            FROM activitydetails ad
            LEFT JOIN items i ON ad.ItemId = i.Id
            LEFT JOIN users u ON ad.CreatedById = u.Id
            WHERE ad.ActivityId = %s
            ORDER BY ad.No ASC
        """
        cur.execute(query, [activity_id])
        details = cur.fetchall()
        cur.close()
        
        for row in details:
            for key, value in row.items():
                if isinstance(value, decimal.Decimal):
                    row[key] = str(value)

        return jsonify({"status": "success", "data": details})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
        
@api_bp.route('/data/<data_type>/<item_id>/toggle-active', methods=['PUT'])
@login_required
@permission_required('SA')
def toggle_active_status(data_type, item_id):
    table_map = {'items': 'items', 'materials': 'materials', 'sizes': 'sizes', 'brands': 'brands', 'models': 'models'}
    table_name = table_map.get(data_type)
    if not table_name:
        return jsonify({"status": "error", "message": "Tipe data tidak valid"}), 400
    try:
        cur = mysql.connection.cursor()
        cur.execute(f"UPDATE {table_name} SET IsActive = CASE WHEN IsActive = 1 THEN 0 ELSE 1 END WHERE Id = %s", [item_id])
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": "Status berhasil diubah."})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/data/items/<item_id>', methods=['GET'])
@login_required
@permission_required('RI')
def get_item_details(item_id):
    try:
        cur = mysql.connection.cursor()
        query = """
            SELECT 
                i.Id, i.Barcode, i.Name, i.Serial, 
                m.Name as MaterialName, s.Name as SizeName, 
                b.Name as BrandName, mo.Name as ModelName, 
                i.Price, i.IsActive, i.CreatedDate, 
                i.MaterialId, i.SizeId, i.BrandId, i.ModelId 
            FROM items i 
            LEFT JOIN materials m ON i.MaterialId = m.Id 
            LEFT JOIN sizes s ON i.SizeId = s.Id 
            LEFT JOIN brands b ON i.BrandId = b.Id 
            LEFT JOIN models mo ON i.ModelId = mo.Id
            WHERE i.Id = %s
        """
        cur.execute(query, [item_id])
        item = cur.fetchone()
        cur.close()
        if not item: return jsonify({"status": "error", "message": "Item tidak ditemukan"}), 404
        
        for key, value in item.items():
            if isinstance(value, bytes):
                # Konversi khusus untuk kolom IsActive (TINYINT)
                if key == 'IsActive':
                    item[key] = 1 if value == b'\x01' else 0
                else:
                    item[key] = value.decode('utf-8')
            elif isinstance(value, decimal.Decimal):
                item[key] = str(value)
            elif isinstance(value, datetime):
                item[key] = value.strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({"status": "success", "data": item})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/data/items/<item_id>', methods=['PUT'])
@login_required
@permission_required('WI')
def update_item(item_id):
    data = request.get_json()
    client_info = data.get('clientInfo', 'Unknown Client')
    name = data.get('Name')
    price = data.get('Price')
    serial = data.get('Serial')
    material_id = data.get('MaterialId')
    size_id = data.get('SizeId')
    brand_id = data.get('BrandId')
    model_id = data.get('ModelId')

    if not all([name, price is not None, material_id, size_id, brand_id, model_id]):
        return jsonify({"status": "error", "message": "Semua field harus diisi"}), 400
    try:
        now, user_id = datetime.now(), session['user_id']
        current_time = now.strftime('%Y-%m-%d %H:%M:%S')
        cur = mysql.connection.cursor()
        cur.execute("UPDATE items SET Name = %s, Price = %s, Serial = %s, MaterialId = %s, SizeId = %s, BrandId = %s, ModelId = %s, ModifiedById = %s, ModifiedDate = %s, ModifiedAt = %s WHERE Id = %s", 
                    (name, price, serial, material_id, size_id, brand_id, model_id, user_id, current_time, client_info, item_id))
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": "Barang berhasil diperbarui."})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/data/<data_type>/<item_id>', methods=['DELETE'])
@login_required
def delete_data(data_type, item_id):
    perm_code = permission_map.get(data_type, {}).get('D')
    if not perm_code or perm_code not in session.get('permissions', []):
        return jsonify({"status": "error", "message": "Akses ditolak."}), 403
    table_name = data_type
    try:
        cur = mysql.connection.cursor()
        cur.execute(f"DELETE FROM {table_name} WHERE Id = %s", [item_id])
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": "Data berhasil dihapus."})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/data/lookup/<data_type>/<item_id>', methods=['PUT'])
@login_required
def update_lookup_data(data_type, item_id):
    perm_code = permission_map.get(data_type, {}).get('W')
    if not perm_code or perm_code not in session.get('permissions', []):
        return jsonify({"status": "error", "message": "Akses ditolak."}), 403
    table_name = data_type
    data = request.get_json()
    new_name = data.get('Name')
    client_info = data.get('clientInfo', 'Unknown Client')
    if not new_name: return jsonify({"status": "error", "message": "Nama tidak boleh kosong"}), 400
    try:
        now, user_id = datetime.now(), session['user_id']
        current_time = now.strftime('%Y-%m-%d %H:%M:%S')
        cur = mysql.connection.cursor()
        cur.execute(f"UPDATE {table_name} SET Name = %s, ModifiedById = %s, ModifiedDate = %s, ModifiedAt = %s WHERE Id = %s", (new_name, user_id, current_time, client_info, item_id))
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": "Data berhasil diperbarui."})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/form-data', methods=['GET'])
@login_required
def get_form_data():
    is_super_admin = 'SA' in session.get('permissions', [])
    try:
        cur = mysql.connection.cursor()
        
        where_clause = "" if is_super_admin else " WHERE IsActive = 1"

        cur.execute(f"SELECT Id, Name FROM sizes{where_clause} ORDER BY Name")
        sizes = cur.fetchall()
        cur.execute(f"SELECT Id, Name FROM brands{where_clause} ORDER BY Name")
        brands = cur.fetchall()
        cur.execute(f"SELECT Id, Name FROM models{where_clause} ORDER BY Name")
        models = cur.fetchall()
        cur.execute(f"SELECT Id, Name FROM materials{where_clause} ORDER BY Name")
        materials = cur.fetchall()
        cur.close()
        
        for item_list in [sizes, brands, models, materials]:
            for item in item_list:
                if isinstance(item.get('Id'), bytes):
                    item['Id'] = item['Id'].decode('utf-8')

        return jsonify({"status": "success", "data": {"sizes": sizes, "brands": brands, "models": models, "materials": materials}})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/add-item', methods=['POST'])
@login_required
@permission_required('WNB') 
def add_item():
    data = request.get_json()
    client_info = data.get('clientInfo', 'Unknown Client')

    # --- 2. Generate data item baru (barcode, id, dll) ---
    now = datetime.now()
    base_10_digits = now.strftime('%y%m%d%H%S')
    milli_2_digits = str(now.microsecond // 10000).zfill(2)
    base_12_digits = base_10_digits + milli_2_digits
    check_digit = calculate_ean13_checksum(base_12_digits)
    barcode = base_12_digits + check_digit
    item_id, user_id = str(uuid.uuid4()), session['user_id']
    current_time = now.strftime('%Y-%m-%d %H:%M:%S')

    cur = None # Inisialisasi cursor di luar try-except
    try:
        cur = mysql.connection.cursor()
        conn_sql = get_sql_server_connection()
        if not conn_sql:
            raise Exception("Gagal terhubung ke database Packing List.")

        # --- 3. Ambil nama 'Size' dari database berdasarkan 'sizeId' ---
        size_id = data.get('sizeId')
        cur.execute("SELECT Name FROM sizes WHERE Id = %s", [size_id])
        size_record = cur.fetchone()
        if not size_record:
            return jsonify({"status": "error", "message": "ID Ukuran tidak valid."}), 400
        size_name = size_record['Name']
        # --- PERUBAHAN: Lakukan INSERT langsung ke SQL Server ---
        cursor_sql = conn_sql.cursor()
        sql_insert = "INSERT INTO dbo.ViewBarcodeSystemCMT (Barcode, Color, Item, Size) VALUES (?, ?, ?, ?);"
        cursor_sql.execute(sql_insert, barcode, '-', data.get('name'), size_name)
        # --- AKHIR PERUBAHAN ---

        # --- 6. Jika berhasil, lanjutkan penyimpanan ke database MySQL lokal ---
        new_item = { 
            'Id': item_id, 'Barcode': barcode, 'Name': data.get('name'), 
            'MaterialId': data.get('materialId'), 'SizeId': data.get('sizeId'), 'IsActive': 1, 
            'CreatedById': user_id, 'CreatedDate': current_time, 'ModifiedById': user_id, 
            'ModifiedDate': current_time, 'BrandId': data.get('brandId'), 'ModelId': data.get('modelId'), 
            'Price': data.get('price', 0.00), 'MinimumStock': 5, 
            'CreatedAt': client_info, 'ModifiedAt': client_info,
            'Serial': data.get('serial') 
        }
        query = "INSERT INTO items (Id, Barcode, Name, MaterialId, SizeId, IsActive, CreatedById, CreatedDate, ModifiedById, ModifiedDate, BrandId, ModelId, Price, MinimumStock, CreatedAt, ModifiedAt, Serial) VALUES (%(Id)s, %(Barcode)s, %(Name)s, %(MaterialId)s, %(SizeId)s, %(IsActive)s, %(CreatedById)s, %(CreatedDate)s, %(ModifiedById)s, %(ModifiedDate)s, %(BrandId)s, %(ModelId)s, %(Price)s, %(MinimumStock)s, %(CreatedAt)s, %(ModifiedAt)s, %(Serial)s)"
        
        cur.execute(query, new_item)
        conn_sql.commit()
        mysql.connection.commit()
        
        return jsonify({"status": "success", "message": f"Item '{data.get('name')}' berhasil ditambahkan."}), 201

    except Exception as e:
        if conn_sql: conn_sql.rollback()
        if mysql.connection: mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        if conn_sql: conn_sql.close()
        if cur: cur.close()

@api_bp.route('/data/lookup/<data_type>/<item_id>', methods=['GET'])
@login_required
def get_lookup_item_details(data_type, item_id):
    # Validasi untuk keamanan
    table_map = {'materials': 'materials', 'sizes': 'sizes', 'brands': 'brands', 'models': 'models'}
    table_name = table_map.get(data_type)
    if not table_name:
        return jsonify({"status": "error", "message": "Tipe data tidak valid"}), 400
    
    # Cek izin baca
    perm_code = permission_map.get(data_type, {}).get('R')
    if not perm_code or perm_code not in session.get('permissions', []):
        return jsonify({"status": "error", "message": "Akses ditolak."}), 403

    try:
        cur = mysql.connection.cursor()
        cur.execute(f"SELECT Id, Name, IsActive FROM {table_name} WHERE Id = %s", [item_id])
        item = cur.fetchone()
        cur.close()
        if not item:
            return jsonify({"status": "error", "message": "Item tidak ditemukan"}), 404
        
        if isinstance(item.get('IsActive'), bytes):
            item['IsActive'] = 1 if item['IsActive'] == b'\x01' else 0
            
        return jsonify({"status": "success", "data": item})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/add-lookup-item', methods=['POST'])
@login_required
def add_lookup_item():
    data = request.get_json()
    client_info = data.get('clientInfo', 'Unknown Client')
    item_type, item_name = data.get('itemType'), data.get('itemName')
    perm_code = permission_map.get(item_type + 's', {}).get('W')
    if not perm_code or perm_code not in session.get('permissions', []):
        return jsonify({"status": "error", "message": "Akses ditolak."}), 403
    table_name = item_type + 's'
    now, user_id = datetime.now(), session['user_id']
    new_id = str(uuid.uuid4())
    current_time = now.strftime('%Y-%m-%d %H:%M:%S')
    new_lookup_item = { 'Id': new_id, 'Name': item_name, 'IsActive': 1, 'CreatedById': user_id, 'CreatedDate': current_time, 'ModifiedById': user_id, 'ModifiedDate': current_time, 'CreatedAt': client_info, 'ModifiedAt': client_info }
    query = f"INSERT INTO {table_name} (Id, Name, IsActive, CreatedById, CreatedDate, ModifiedById, ModifiedDate, CreatedAt, ModifiedAt) VALUES (%(Id)s, %(Name)s, %(IsActive)s, %(CreatedById)s, %(CreatedDate)s, %(ModifiedById)s, %(ModifiedDate)s, %(CreatedAt)s, %(ModifiedAt)s)"
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, new_lookup_item)
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": f"{item_name} berhasil ditambahkan.", "newItem": {"Id": new_id, "Name": item_name}}), 201
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/pajak-invoice/stores', methods=['GET'])
@login_required
@permission_required('RPjk')
def get_pajak_invoice_stores():
    """Mengambil daftar unik toko dari view pajak invoice."""
    conn_sql = get_jbbdata_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database JBBData."}), 503
    try:
        cursor = conn_sql.cursor()
        query = "SELECT DISTINCT Toko FROM dbo.ViewSalInvDetPajakExt WHERE Toko IS NOT NULL AND Toko <> '' ORDER BY Toko;"
        cursor.execute(query)
        stores = [row.Toko for row in cursor.fetchall()]
        return jsonify({"status": "success", "data": stores})
    except pyodbc.Error as ex:
        return jsonify({"status": "error", "message": f"Database error: {str(ex)}"}), 500
    finally:
        if conn_sql:
            conn_sql.close()

@api_bp.route('/pajak-invoice/export', methods=['GET'])
@login_required
@permission_required('RPjk')
def export_pajak_invoice():
    """Mengekspor data pajak invoice ke file Excel berdasarkan filter."""
    search_query = request.args.get('search', '')
    toko = request.args.get('toko')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn_sql = get_jbbdata_sql_server_connection()
    if not conn_sql:
        return "Gagal terhubung ke database.", 503

    try:
        cursor = conn_sql.cursor()
        params = []
        
        query_base = "FROM dbo.ViewSalInvDetPajakExt"
        query_fields = "InvNo, CONVERT(VARCHAR, InvDate, 105) as InvDate, Toko, Customer, Qty, Up, Amount, Disc, TotalAmount, DppUp, DppDisc, DppAmt, DppNilaiLain, VatAmount, Remark"
        
        where_clauses = []
        if search_query:
            where_clauses.append("(InvNo LIKE ? OR Toko LIKE ? OR Customer LIKE ?)")
            search_term = f'%{search_query}%'
            params.extend([search_term, search_term, search_term])
        
        if toko and toko != 'All':
            where_clauses.append("Toko = ?")
            params.append(toko)

        if start_date and end_date:
            where_clauses.append("CAST(InvDate AS DATE) BETWEEN ? AND ?")
            params.extend([start_date, end_date])

        where_clause_str = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
        
        query = f"SELECT {query_fields} {query_base} {where_clause_str} ORDER BY InvDate DESC, InvNo ASC"
        cursor.execute(query, params)
        
        rows = cursor.fetchall()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Pajak Invoice Penjualan"
        
        headers = ['No. Invoice', 'Tgl. Invoice', 'Toko', 'Customer', 'Qty', 'Harga Satuan', 'Jumlah', 'Diskon', 'Total', 'DPP Harga', 'DPP Diskon', 'DPP Jumlah', 'DPP Lain', 'PPN', 'Keterangan']
        sheet.append(headers)
        
        for row in rows:
            processed_row = [float(item) if isinstance(item, decimal.Decimal) else item for item in row]
            sheet.append(processed_row)

        # --- Logic to create dynamic filename ---
        filename_parts = ["Pajak_Invoice_Penjualan"]
        
        # Add store suffix
        if toko and toko != 'All':
            # Replace spaces with underscores for better compatibility
            safe_toko_name = toko.replace(' ', '_')
            filename_parts.append(safe_toko_name)

        # Add month and year suffix
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                month_names_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
                month_name = month_names_id[date_obj.month - 1]
                year = date_obj.year
                filename_parts.extend([month_name, str(year)])
            except (ValueError, IndexError):
                pass # Ignore if date is invalid
        
        download_name = "_".join(filename_parts) + ".xlsx"
        # --- End of logic ---

        mem_file = io.BytesIO()
        workbook.save(mem_file)
        mem_file.seek(0)
        
        return send_file(mem_file, as_attachment=True, download_name=download_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except pyodbc.Error as ex:
        return f"Database error: {str(ex)}", 500
    finally:
        if conn_sql:
            conn_sql.close()

# --- ENDPOINT ANALYTICS DIPERBARUI ---
# --- ENDPOINT ANALYTICS DIPERBARUI ---
@api_bp.route('/analytics/daily')
@login_required
@permission_required('RA')
def get_analytics_daily():
    target_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"status": "error", "message": "Format tanggal tidak valid."}), 400

    try:
        cur = mysql.connection.cursor()
        
        # Query untuk total penjualan hari ini dengan rincian
        total_query = """
            SELECT 
                SUM(TotalPrice) as total_penjualan,
                SUM(Total) as total_lusin,
                SUM(CASE WHEN PaymentType IN ('EDCDebit', 'Cash', 'Transfer') THEN TotalPrice ELSE 0 END) as lunas,
                SUM(CASE WHEN PaymentType = 'Debt' THEN TotalPrice ELSE 0 END) as belum_lunas,
                SUM(CASE WHEN PaymentType IS NULL OR PaymentType = '' THEN TotalPrice ELSE 0 END) as tidak_diketahui
            FROM activities 
            WHERE Type = 'SalesOrder' AND DATE(CreatedDate) = %s
        """
        cur.execute(total_query, [target_date])
        total_result = cur.fetchone()

        # Query untuk penjualan terkini
        recent_query = """
            SELECT 
                a.Name, a.TotalPrice, TIME_FORMAT(a.CreatedDate, '%%H:%%i:%%s') as Waktu, b.Name as Cabang 
            FROM activities a 
            LEFT JOIN branches b ON a.BranchId = b.Id 
            WHERE a.Type = 'SalesOrder' AND DATE(a.CreatedDate) = %s 
            ORDER BY a.CreatedDate DESC 
        """
        cur.execute(recent_query, [target_date])
        recent_results = cur.fetchall()
        
        cur.close()

        # Format semua data Decimal ke string
        if total_result:
            for key, value in total_result.items():
                if isinstance(value, decimal.Decimal):
                    total_result[key] = str(value)
        
        for row in recent_results:
            if isinstance(row.get('TotalPrice'), decimal.Decimal):
                row['TotalPrice'] = str(row['TotalPrice'])

        return jsonify({
            "status": "success", 
            "data": { "total": total_result, "recent": recent_results }
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# --- ENDPOINT BARU UNTUK REKAP BULANAN ---
@api_bp.route('/analytics/monthly')
@login_required
@permission_required('RA')
def get_analytics_monthly():
    try:
        target_month = date.today().month
        target_year = date.today().year

        cur = mysql.connection.cursor()
        query = """
            SELECT b.Name as Cabang, s.Name as Salesman, SUM(a.TotalPrice) as TotalPenjualan
            FROM activities a
            LEFT JOIN branches b ON a.BranchId = b.Id
            LEFT JOIN salesmans s ON a.SalesmanId = s.Id
            WHERE a.Type = 'SalesOrder' AND MONTH(a.CreatedDate) = %s AND YEAR(a.CreatedDate) = %s
            GROUP BY b.Name, s.Name
            ORDER BY b.Name, TotalPenjualan DESC
        """
        cur.execute(query, [target_month, target_year])
        results = cur.fetchall()
        cur.close()

        for row in results:
            if isinstance(row.get('TotalPenjualan'), decimal.Decimal):
                row['TotalPenjualan'] = str(row['TotalPenjualan'])
        
        return jsonify({"status": "success", "data": results})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
# Hapus fungsi get_analytics_daily dan get_analytics_monthly yang lama

# --- FUNGSI ANALYTICS YANG SUDAH DIPERBARUI ---
@api_bp.route('/analytics/range')
@login_required
@permission_required('RA')
def get_analytics_by_range():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({"status": "error", "message": "Parameter start_date dan end_date diperlukan."}), 400

    try:
        datetime.strptime(start_date_str, '%Y-%m-%d')
        datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({"status": "error", "message": "Format tanggal tidak valid. Gunakan YYYY-MM-DD."}), 400

    try:
        cur = mysql.connection.cursor()
        
        total_query = """
            SELECT 
                SUM(TotalPrice) as total_penjualan,
                SUM(Total) as total_lusin,
                SUM(CASE WHEN PaymentType IN ('EDCDebit', 'Cash', 'Transfer') THEN TotalPrice ELSE 0 END) as lunas,
                SUM(CASE WHEN PaymentType = 'Debt' THEN TotalPrice ELSE 0 END) as belum_lunas,
                SUM(CASE WHEN PaymentType IS NULL OR PaymentType = '' THEN TotalPrice ELSE 0 END) as tidak_diketahui
            FROM activities 
            WHERE Type = 'SalesOrder' AND DATE(CreatedDate) BETWEEN %s AND %s
        """
        cur.execute(total_query, [start_date_str, end_date_str])
        total_result = cur.fetchone()

        recent_query = """
            SELECT 
                a.Id, a.Name, a.TotalPrice, DATE_FORMAT(a.CreatedDate, '%%d-%%m-%%Y %%H:%%i') as Waktu, b.Name as Cabang 
            FROM activities a 
            LEFT JOIN branches b ON a.BranchId = b.Id 
            WHERE a.Type = 'SalesOrder' AND DATE(a.CreatedDate) BETWEEN %s AND %s 
            ORDER BY a.CreatedDate DESC
        """
        cur.execute(recent_query, [start_date_str, end_date_str])
        recent_results = cur.fetchall()
        
        cur.close()

        if total_result:
            for key, value in total_result.items():
                if isinstance(value, decimal.Decimal):
                    total_result[key] = str(value)
        
        for row in recent_results:
            if isinstance(row.get('TotalPrice'), decimal.Decimal):
                row['TotalPrice'] = str(row['TotalPrice'])

        return jsonify({
            "status": "success", 
            "data": { "total": total_result, "recent": recent_results }
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
@api_bp.route('/analytics/timeseries')
@login_required
@permission_required('RA')
def get_analytics_timeseries():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({"status": "error", "message": "Parameter start_date dan end_date diperlukan."}), 400

    try:
        # Validasi format tanggal
        datetime.strptime(start_date_str, '%Y-%m-%d')
        datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({"status": "error", "message": "Format tanggal tidak valid. Gunakan YYYY-MM-DD."}), 400

    try:
        cur = mysql.connection.cursor()
        query = """
            SELECT 
                DATE(CreatedDate) as tanggal, 
                SUM(TotalPrice) as total_penjualan
            FROM activities
            WHERE Type = 'SalesOrder' AND DATE(CreatedDate) BETWEEN %s AND %s
            GROUP BY DATE(CreatedDate)
            ORDER BY DATE(CreatedDate) ASC;
        """
        cur.execute(query, [start_date_str, end_date_str])
        results = cur.fetchall()
        cur.close()

        for row in results:
            if isinstance(row.get('total_penjualan'), decimal.Decimal):
                row['total_penjualan'] = str(row['total_penjualan'])
            if isinstance(row.get('tanggal'), date):
                row['tanggal'] = row['tanggal'].strftime('%Y-%m-%d')

        return jsonify({"status": "success", "data": results})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    
@api_bp.route('/packinglist/upload-csv', methods=['POST'])
@login_required
@permission_required('WPL')
def upload_packing_list_csv():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah."}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Tidak ada file yang dipilih."}), 400

    if file and file.filename.endswith('.csv'):
        try:
            # Mengubah file stream menjadi format teks yang bisa dibaca
            file_text = io.TextIOWrapper(file, encoding='utf-8')
            
            # Membaca CSV sebagai dictionary
            reader = csv.DictReader(file_text)
            
            # Validasi header
            required_headers = {'Barcode', 'Color', 'Item', 'Size'}
            if not required_headers.issubset(reader.fieldnames):
                return jsonify({"status": "error", "message": f"File CSV harus memiliki header: {', '.join(required_headers)}"}), 400
            
            conn_sql = get_sql_server_connection()
            if not conn_sql:
                raise Exception("Gagal terhubung ke database Packing List.")
            
            cursor = conn_sql.cursor()
            inserted_count = 0
            updated_count = 0
            total_records_received = 0
            # --- PERUBAHAN: Logika UPSERT langsung ke DB ---
            sql_check = "SELECT 1 FROM dbo.ViewBarcodeSystemCMT WHERE Barcode = ?;"
            sql_update = "UPDATE dbo.ViewBarcodeSystemCMT SET Color = ?, Item = ?, Size = ? WHERE Barcode = ?;"
            sql_insert = "INSERT INTO dbo.ViewBarcodeSystemCMT (Barcode, Color, Item, Size) VALUES (?, ?, ?, ?);"

            for row in reader:
                total_records_received += 1
                barcode = row.get('Barcode')
                if not barcode: continue

                cursor.execute(sql_check, barcode)
                if cursor.fetchone():
                    cursor.execute(sql_update, row.get('Color', ''), row.get('Item', ''), row.get('Size', ''), barcode)
                    updated_count += 1
                else:
                    cursor.execute(sql_insert, barcode, row.get('Color', ''), row.get('Item', ''), row.get('Size', ''))
                    inserted_count += 1
            
            conn_sql.commit()
            # --- AKHIR PERUBAHAN ---
            
            return jsonify({
                "message": "Data berhasil diproses.",
                "total_records_received": total_records_received,
                "total_records_inserted": inserted_count,
                "total_records_updated": updated_count
            }), 200
        except Exception as e:
            return jsonify({"status": "error", "message": f"Gagal memproses file CSV: {e}"}), 500
        finally:
         if 'conn_sql' in locals() and conn_sql: conn_sql.close()

# ... (di dalam file api.py)

@api_bp.route('/packinglist/delete/<string:barcode>', methods=['DELETE'])
@login_required
@permission_required('WPL')
def delete_packing_list_item(barcode):
    conn_sql = get_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database Packing List."}), 503
    
    try:
        cursor = conn_sql.cursor()
        # --- PERUBAHAN: DELETE langsung dari DB ---
        cursor.execute("DELETE FROM dbo.ViewBarcodeSystemCMT WHERE Barcode = ?", barcode)
        conn_sql.commit()
        # --- AKHIR PERUBAHAN ---
        
        if cursor.rowcount > 0:
            return jsonify({"status": "success", "message": f"Barcode {barcode} berhasil dihapus."}), 200
        else:
            return jsonify({"status": "error", "message": "Barcode tidak ditemukan."}), 404
            
    except pyodbc.Error as ex:
        conn_sql.rollback()
        return jsonify({"status": "error", "message": f"Database error: {str(ex)}"}), 500
    finally:
        conn_sql.close()


@api_bp.route('/packinglist/detail/<int:idno>')
@login_required
@permission_required('RPL')
def get_packing_list_detail(idno):
    conn_sql = get_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database Packing List."}), 503
    
    try:
        cursor = conn_sql.cursor()
        # --- PERUBAHAN: SELECT langsung dari DB ---
        query = "SELECT InvId, Size, Color, Qty FROM dbo.PackListCmtDet WHERE IDNo = ?;"
        cursor.execute(query, idno)
        # --- AKHIR PERUBAHAN ---
        
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return jsonify({"status": "success", "data": results}), 200
            
    except pyodbc.Error as ex:
        return jsonify({"status": "error", "message": f"Database error: {str(ex)}"}), 500
    finally:
        conn_sql.close()
    


# --- TAMBAHKAN FUNGSI BARU UNTUK MEMBACA HEADER EXCEL ---
def clean_headers(headers):
    cleaned_headers = []
    unnamed_count = 1
    for header in headers:
        if header is None or str(header).strip() == '':
            cleaned_headers.append(f"Kolom Tanpa Nama {unnamed_count}")
            unnamed_count += 1
        else:
            cleaned_headers.append(str(header).strip())
    return cleaned_headers


@api_bp.route('/multipayroll/get-headers', methods=['POST'])
@login_required
@permission_required('RA')
def get_xlsx_headers():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah."}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "Format file harus .xlsx"}), 400
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        # --- PERBAIKAN DI SINI ---
        cleaned_headers = clean_headers(headers)
        # --- AKHIR PERBAIKAN --
        return jsonify({"status": "success", "headers": cleaned_headers})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Gagal membaca header file: {str(e)}"}), 500

# --- GANTI TOTAL FUNGSI CONVERT DENGAN YANG BARU INI ---
@api_bp.route('/multipayroll/convert', methods=['POST'])
@login_required
@permission_required('RA')
def convert_multipayroll_file():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah."}), 400

    file = request.files['file']
    try:
        # Ambil semua data dari form
        settings = json.loads(request.form.get('settings'))
        mapping = json.loads(request.form.get('mapping'))

        # Baca file Excel
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

         # --- PERBAIKAN DI SINI ---
        cleaned_headers = clean_headers(headers)
        # --- AKHIR PERBAIKAN ---

        # Buat map dari nama kolom ke indeksnya untuk pencarian cepat
        header_map = {header: i for i, header in enumerate(cleaned_headers)}

        detail_lines = []
        total_amount = 0.0
        total_records = 0

        # Iterasi baris data (mulai dari baris ke-2)
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue # Lewati baris kosong

            detail_data = {
                'Transaction ID': '', 'Transfer Type': 'BCA', 'Beneficiary ID': '',
                'Credited Account': '', 'Receiver Name': '', 'Amount': '0.00',
                'Employee ID': '', 'Remark': 'Gaji', 'Email': '', 'SWIFT': '',
                'Cust Type': '', 'Cust Residence': ''
            }

            # Isi detail_data menggunakan mapping dari frontend
            for field, column_name in mapping.items():
                if column_name and column_name in header_map:
                    cell_value = row[header_map[column_name]]
                    detail_data[field] = str(cell_value).strip() if cell_value is not None else ''

            # Validasi Amount
            try:
                amount = float(detail_data['Amount'])
                total_amount += amount
                detail_data['Amount'] = f"{amount:.2f}"
            except (ValueError, TypeError):
                return jsonify({"status": "error", "message": f"Nilai 'Amount' tidak valid di baris {row_index}."}), 400

            # Buat baris detail
            line = "1|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}".format(
                f"000000000{detail_data['Transaction ID']}", detail_data['Transfer Type'],
                detail_data['Beneficiary ID'], detail_data['Credited Account'],
                detail_data['Receiver Name'], detail_data['Amount'],
                detail_data['Employee ID'], detail_data['Remark'],
                detail_data['Email'], detail_data['SWIFT'],
                detail_data['Cust Type'], detail_data['Cust Residence']
            )
            detail_lines.append(line)
            total_records += 1

        # Buat baris header
        header_line = "0|PY|{}|{}|{}|{}|{}|{}|{}|||".format(
            settings['corporate_id'], 
            settings['company_code'],
            # --- PERBAIKAN DI DUA BARIS INI ---
            f"{int(settings['file_no']):08d}", 
            settings['transfer_date'],
            f"{int(settings['transfer_time']):02d}", 
            # --- AKHIR PERBAIKAN ---
            settings['source_account'],
            f"{total_records:05d}"
        )

        payroll_content = header_line + "\n" + "\n".join(detail_lines)

        return jsonify({
            "status": "success",
            "payroll_data": payroll_content,
            "total_records": total_records,
            "total_amount": f"{total_amount:.2f}"
        })

    except Exception as e:
        return jsonify({"status": "error", "message": f"Terjadi kesalahan: {str(e)}"}), 500
    
@api_bp.route('/setting/<key>', methods=['GET'])
@login_required
@permission_required('RA') # Izin untuk membaca pengaturan
def get_setting(key):
    try:
        cur = mysql.connection.cursor()
        cur.execute("SELECT setting_value FROM app_settings WHERE setting_key = %s", [key])
        result = cur.fetchone()
        cur.close()
        if result:
            return jsonify({"status": "success", "value": result['setting_value']})
        else:
            # Jika key tidak ditemukan, kembalikan nilai default atau error
            return jsonify({"status": "error", "message": "Setting tidak ditemukan."}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@api_bp.route('/setting', methods=['POST'])
@login_required
@permission_required('RA') # Izin untuk menulis pengaturan
def update_setting():
    data = request.get_json()
    key = data.get('key')
    value = data.get('value')

    if not key or value is None:
        return jsonify({"status": "error", "message": "Key dan value wajib diisi."}), 400

    try:
        cur = mysql.connection.cursor()
        # Gunakan INSERT ... ON DUPLICATE KEY UPDATE untuk UPSERT
        cur.execute("""
            INSERT INTO app_settings (setting_key, setting_value) 
            VALUES (%s, %s) 
            ON DUPLICATE KEY UPDATE setting_value = %s
        """, (key, value, value))
        mysql.connection.commit()
        cur.close()
        return jsonify({"status": "success", "message": "Setting berhasil diperbarui."})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500



CHAR_TABLE = "XRVZK2TS7QFG0CHELA4OPM9IDYUWJNBo5cx1t3hwry8knabq7efuvijmlpsgzd!@#$%&*()-+=\\:;\"<>,.?/ '"
INT_ADDER = 3371
INT_SEED = 3751517

def mtdCalculateAutoCollect(strInput):
    iCheckSum = 0
    for index1 in range(1, len(strInput) + 1):
        x = 0
        for index2 in range(1, len(CHAR_TABLE) + 1):
            if ord(CHAR_TABLE[index2 - 1]) == ord(strInput[index1 - 1]):
                num2 = index1 + index2
                x = num2 % INT_ADDER
                iCheckSum += (num2 + INT_ADDER * (INT_ADDER + index1) + INT_ADDER * (INT_ADDER + index2)) + pow(x, 2)
            else:
                iCheckSum += INT_ADDER * index1 + pow(x, 2)
    return iCheckSum

def mtdGenerateH2HChecksum(lines):
    iCheckSum = 0
    for line in lines:
        iCheckSum += mtdCalculateAutoCollect(line.strip())
    iCheckSum += INT_SEED
    return str(iCheckSum)

@api_bp.route('/multipayroll/calculate-checksum', methods=['POST'])
@login_required
@permission_required('RA')
def calculate_payroll_checksum():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang diunggah."}), 400

    file = request.files['file']
    if not file.filename.endswith('.txt'):
        return jsonify({"status": "error", "message": "Format file harus .txt"}), 400

    try:
        content = file.read().decode('utf-8')
        lines = content.splitlines()
        checksum = mtdGenerateH2HChecksum(lines)
        return jsonify({"status": "success", "checksum": checksum})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Gagal menghitung checksum: {str(e)}"}), 500
    
# ... (di dalam file api.py)


def get_jbbdata_sql_server_connection():
    """Membuat koneksi ke database SQL Server untuk JBBData."""
    try:
        conn = pyodbc.connect(
            'DRIVER={' + os.getenv('SQL_SERVER_DRIVER') + '};'
            'SERVER=' + os.getenv('SQL_SERVER_HOST') + ';'
            'DATABASE=' + os.getenv('JBBData_DATABASE') + ';'
            'UID=' + os.getenv('SQL_SERVER_USER') + ';'
            'PWD=' + os.getenv('SQL_SERVER_PASSWORD') + ';'
        )
        return conn
    except pyodbc.Error as ex:
        print(f"Gagal terhubung ke SQL Server (JBBData): {ex}")
        return None
@api_bp.route('/analytics/factory-range')
@login_required
@permission_required('RA')
def get_factory_sales_by_range():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({"status": "error", "message": "Parameter tanggal diperlukan."}), 400

    conn_sql = get_jbbdata_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database Pabrik."}), 503

    try:
        cursor = conn_sql.cursor()
        query = """
            SELECT 
                SUM(Amount) as total_penjualan,
                SUM(Qty) as total_lusin
            FROM dbo.ViewSalesReportExt 
            WHERE CAST(InvDate AS DATE) BETWEEN ? AND ?
        """
        cursor.execute(query, start_date_str, end_date_str)
        result = cursor.fetchone()
        
        # Ambil 10 penjualan terbaru
        recent_query = """
            WITH InvoiceTotals AS (
                SELECT
                    InvNo,
                    Customer,
                    MIN(InvDate) as InvDate,
                    SUM(Amount) as TotalAmount
                FROM dbo.ViewSalesReportExt
                WHERE CAST(InvDate AS DATE) BETWEEN ? AND ?
                GROUP BY InvNo, Customer
            )
            SELECT TOP 10
                InvNo,
                CONVERT(VARCHAR, InvDate, 105) + ' ' + LEFT(CONVERT(VARCHAR, InvDate, 108), 5) as Waktu,
                Customer,
                TotalAmount as Amount
            FROM InvoiceTotals
            ORDER BY InvDate DESC
        """
        cursor.execute(recent_query, start_date_str, end_date_str)
        recent_results = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
        
        # --- PERBAIKAN UTAMA DI SINI ---
        # Periksa apakah query mengembalikan hasil. Jika tidak, artinya tidak ada penjualan.
        if result and result.total_penjualan is not None:
            # Jika ada penjualan, proses seperti biasa
            summary = {
                'total_penjualan': str(result.total_penjualan),
                'total_lusin': str(result.total_lusin or 0)
            }
        else:
            # Jika tidak ada penjualan, kembalikan nilai nol
            summary = {
                'total_penjualan': '0',
                'total_lusin': '0'
            }
        # --- AKHIR PERBAIKAN ---
        
        # Format desimal di data penjualan terbaru
        for row in recent_results:
            if isinstance(row.get('Amount'), decimal.Decimal):
                row['Amount'] = str(row['Amount'])

        return jsonify({"status": "success", "data": {"total": summary, "recent": recent_results}})

    except Exception as e: # Tangkap semua jenis error untuk penanganan yang lebih baik
        return jsonify({"status": "error", "message": f"Terjadi kesalahan internal: {str(e)}"}), 500
    finally:
        if conn_sql:
            conn_sql.close()

@api_bp.route('/factory-sales/detail/<inv_no>')
@login_required
@permission_required('RA')
def get_factory_sale_detail(inv_no):
    conn_sql = get_jbbdata_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database Pabrik."}), 503
    try:
        cursor = conn_sql.cursor()
        
        # Query untuk detail item
        detail_query = """
            SELECT
                InvName as Name,
                Size,
                Qty,
                Up as HargaQty,
                Amount
            FROM dbo.ViewSalesReportExt
            WHERE InvNo = ?
            ORDER BY InvName, Size
        """
        cursor.execute(detail_query, inv_no)
        details = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
        
        # Format desimal ke string
        for row in details:
            for key, value in row.items():
                if isinstance(value, decimal.Decimal):
                    row[key] = str(value)

        return jsonify({"status": "success", "data": details})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        if conn_sql:
            conn_sql.close()
@api_bp.route('/analytics/factory-timeseries')
@login_required
@permission_required('RA')
def get_factory_analytics_timeseries():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({"status": "error", "message": "Parameter tanggal diperlukan."}), 400

    conn_sql = get_jbbdata_sql_server_connection()
    if not conn_sql:
        return jsonify({"status": "error", "message": "Gagal terhubung ke database Pabrik."}), 503

    try:
        cursor = conn_sql.cursor()
        query = """
            SELECT 
                CAST(InvDate AS DATE) as tanggal, 
                SUM(Amount) as total_penjualan
            FROM dbo.ViewSalesReportExt
            WHERE CAST(InvDate AS DATE) BETWEEN ? AND ?
            GROUP BY CAST(InvDate AS DATE)
            ORDER BY tanggal ASC;
        """
        cursor.execute(query, start_date_str, end_date_str)
        
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Format data agar konsisten
        for row in results:
            if isinstance(row.get('total_penjualan'), decimal.Decimal):
                row['total_penjualan'] = str(row['total_penjualan'])
            if isinstance(row.get('tanggal'), (datetime, date)):
                row['tanggal'] = row['tanggal'].strftime('%Y-%m-%d')

        return jsonify({"status": "success", "data": results})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Gagal mengambil data tren: {str(e)}"}), 500
    finally:
        if conn_sql:
            conn_sql.close()

@api_bp.route('/analytics/timeseries-hourly')
@login_required
@permission_required('RA')
def get_analytics_timeseries_hourly():
    target_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"status": "error", "message": "Format tanggal tidak valid."}), 400

    try:
        cur = mysql.connection.cursor()
        # Query penjualan per jam
        query = """
            SELECT 
                HOUR(CreatedDate) as hour,
                SUM(TotalPrice) as total_penjualan
            FROM activities
            WHERE Type = 'SalesOrder' AND DATE(CreatedDate) = %s
            GROUP BY HOUR(CreatedDate)
            ORDER BY hour ASC;
        """
        cur.execute(query, [target_date])
        results = cur.fetchall()
        cur.close()

        # Format hasil agar jam 0-23 selalu ada
        timeseries = []
        result_map = {row['hour']: str(row['total_penjualan']) if row['total_penjualan'] is not None else '0' for row in results}
        for h in range(24):
            timeseries.append({
                'hour': h,
                'total_penjualan': result_map.get(h, '0')
            })

        return jsonify({"status": "success", "timeseries": timeseries})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500